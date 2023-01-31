import os
import requests
import urllib.request
import xlsxwriter
from bs4 import BeautifulSoup
import webbrowser
import time
import re
import selenium
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.remote.webelement import WebElement
from selenium.common.exceptions import NoSuchAttributeException
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import ElementNotInteractableException
from selenium.common.exceptions import ElementNotVisibleException
from optparse import OptionParser
from concurrent.futures import ThreadPoolExecutor
import pandas as pd
import openpyxl
class SportsOracles():
    def names(self, website):
        result = requests.get(website)
        src = result.content
        soup = BeautifulSoup(src, 'lxml')
        name1 = []
        for span_tag in soup.find_all('span'):
            a_tag = span_tag.find('a')
            try:
                if 'title' in a_tag.attrs:
                    temp = str(a_tag)
                    start = temp.find('>') + 1
                    end = temp.find('<', start)
                    name = temp[start:end]
                    if name.find(" "):
                        name1.append(name)
            except:
                pass
        return name1
    def madden(self, website):
        result = requests.get(website)
        src = result.content
        soup = BeautifulSoup(src, 'lxml')
        madden1 = []
        substrings = ["diamond", "amethyst", "ruby", "sapphire", "emerald", "gold", 
        "silver", "bronze", "dark-matter", "opal", "pinkdiamond"]
        for td_tag in soup.find_all('td'):
            span_tag = td_tag.find('span')
            try:
                if 'class' in span_tag.attrs:
                    temp = str(span_tag)
                    if any (substring in temp for substring in substrings):
                        start = temp.find('>') + 1
                        end = temp.find('<', start)
                        maddenOvr = temp[start:end]
                        if maddenOvr.isdigit():
                            # print(maddenOvr)
                            madden1.append(maddenOvr)
            except:
                pass
        # print(madden1)
        return madden1
    def Position(self, website):
        result = requests.get(website)
        src = result.content
        soup = BeautifulSoup(src, 'lxml')
        pos1 = []
        for span_tag in soup.find_all('span'):
            a_tag = span_tag.find('a')
            try:
                if 'title' in a_tag.attrs:
                    temp = str(a_tag)
                    start = temp.find('>') + 1
                    end = temp.find('<', start)
                    pos = temp[start:end]
                    if not pos.find(" "):
                        # positions = {"LE": "DE", "RE": "DE", 
                        # "LT": "OT", "RT": "OT", "LG": "OG", "RG": "OG", "ROLB": "OLB", 
                        # "LOLB": "OLB"}
                        # pos2 = positions.get(pos, pos)
                        if "LE" in pos:
                            pos = "DE"
                        elif "RE" in pos:
                            pos = "DE"
                        elif  "LT" in pos:
                            pos = "OT"
                        elif  "RT" in pos:
                            pos = "OT"
                        elif  "LG" in pos:
                            pos = "OG"
                        elif  "RG" in pos:
                            pos = "OG"
                        elif "ROLB" in pos:
                            pos = "OLB"
                        elif "LOLB" in pos:
                            pos = "OLB"
                        # print(pos)
                        pos1.append(pos)
            except:
                pass
        # print(pos1)
        return pos1
    def PFF(self, str, team):
        driver = webdriver.Chrome('./chromedriver')
        url = 'https://www.google.com/'
        driver.get(url)
        element1 = driver.find_element(By.NAME, "q")
        search = "PFF " + str + " " + team
        element1.send_keys(search)
        element1.send_keys(Keys.RETURN)
        link = driver.find_elements(By.CSS_SELECTOR, "a[ping^='/url?sa=t&source=web&rct=j&url=https://www.pff.com/nfl/playe'] > .DKV0Md.LC20lb.MBeuO")
        if not link:
            # print("NOT FOUND **")
            element5 = "N/A"
            element6 = "N/A"
        else:
            element2 = driver.find_element(By.CSS_SELECTOR, "a[ping^='/url?sa=t&source=web&rct=j&url=https://www.pff.com/nfl/playe'] > .DKV0Md.LC20lb.MBeuO").click()
            driver.maximize_window()
            try:
                PFFAge = driver.find_element(By.TAG_NAME, "small")
                element3 = PFFAge.text
                element5 = element3.translate({ord(c): "" for c in ' ()'})
                # print("Age: " + element5)
            except:
                # print("Age Not Found")
                element5 = "N/A"
            try:
                PFFOvrLink = driver.find_element(By.CSS_SELECTOR, ".kyber-grade-badge__info-text")
                element6 = PFFOvrLink.text
                # print("PFF: " + element6)
            except NoSuchElementException:
                # print("Overall Not Found")
                element6 = "N/A"
        return element5, element6 
    def SpotracBackup3(self, str, team): #go to website link when googling since it'd be hard 
        driver = webdriver.Chrome('./chromedriver')
        url = 'https://www.spotrac.com/nfl/'
        driver.get(url)
        driver.maximize_window()
        element1 = driver.find_element(By.NAME, "search")
        search =  str + " average salary spotrac"
        element1.send_keys(search)
        element1.send_keys(Keys.RETURN)
        link = driver.find_elements(By.ClassName, "playerValue")
        if not link:
            element1 = SportsOracles().SpotracBackup1(str, team)
        else:
            element1 = link.text
        # driver = webdriver.Chrome('./chromedriver')
        # url = 'google.com'
        # driver.get(url)
        # driver.maximize_window()
        # element1 = driver.find_element(By.NAME, "q")
        # search =  team + "NFL spotrac contract"
        # element1.send_keys(search)
        # element1.send_keys(Keys.RETURN)
        # link = driver.find_elements(By.CSS_SELECTOR, "div:nth-of-type(2) > .Ww4FFb.g.tF2Cxc.vt6azd  .UK95Uc.Z26q7c.jGGQ5e .DKV0Md.LC20lb.MBeuO")
        # if not link:
        #     element1 = SportsOracles().SpotracBackup2(str, team)
        # else:
        #     element1 = link.text
        # print(element1)
        # return element1
        string = "Didn't Work"
        print(string)
        return string
    def SpotracBackup2(self, str, team):
        # driver = webdriver.Chrome('./chromedriver')
        # url = 'https://www.spotrac.com/nfl/'
        # driver.get(url)
        # driver.maximize_window()
        # element1 = driver.find_element(By.NAME, "search")
        # search = str
        # element1.send_keys(search)
        # element1.send_keys(Keys.RETURN)
        # link = driver.find_elements(By.ClassName, "playerValue")
        # if not link:
        #     element1 = SportsOracles().SpotracBackup1(str, team)
        # else:
        #     element1 = link.text
        string = "Didn't Work"
        print(string)
        return string
    def SpotracBackup1(self, str, team): #just go to cardinals average salary website and look for the player and then the next number
        print("4")
        driver = webdriver.Chrome('./chromedriver')
        url = 'https://www.spotrac.com/nfl/'
        driver.get(url)
        print("5")
        driver.maximize_window()
        print("6")
        element1 = driver.find_element(By.NAME, "search")
        search = str
        element1.send_keys(search)
        element1.send_keys(Keys.RETURN)
        print("7")
        try:
            link = driver.find_elements(By.ClassName, "playerValue").click()
            element1 = link.text
        except:
            element1 = SportsOracles().SpotracBackup2(str, team)
        print("Second: " + element1)
        return element1
    def SpotracBackup(self, str, team):
        print("1")
        driver = webdriver.Chrome('./chromedriver')
        url = 'https://www.google.com/'
        driver.get(url)
        element1 = driver.find_element(By.NAME, "q")
        search =  str + " " + team + " average salary spotrac"
        print("2")
        element1.send_keys(search)
        element1.send_keys(Keys.RETURN)
        print("3")
        link = driver.find_elements(By.CSS_SELECTOR, ".AZCkJd.Z0LcW.d2J77b.t2b5Cf > .IZ6rdc")
        if not link:
            # element1 = SportsOracles().SpotracBackup1(str, team)
            element1 = "N/A"
        else:
            element1 = link[0].text
            if not any(c.isdigit() for c in element1):
                link2 = driver.find_elements(By.CSS_SELECTOR, ".V3FYCf b")
                if not link2:
                    element1 = "N/A"
                else:
                    element1 = link2[0].text
                    words = element1.split()

                        # Initialize a variable to store the index of the word "average"
                    average_index = None
                    try:                         
                        # Iterate through the words in the paragraph
                        for i, word in enumerate(words):
                            if word.lower() == "average":
                                average_index = i
                                break

                        # Initialize a variable to store the dollar amount
                        dollar_amount = None

                        # Iterate through the words after the word "average"
                        for i in range(average_index + 1, len(words)):
                            # Check if the word starts with a "$"
                            if re.match(r"^\$", words[i]):
                                element1 = words[i]
                                break
                            # If the word doesn't start with a "$" or if it contains any letters, break the loop
                            elif not re.match(r"^\$\d+", words[i]) or any(c.isalpha() for c in words[i]):
                                break
                        element1 = element1.replace('average annual salary of ','')
                        print(element1)
                    except:
                        print("bug")
                        element1 = "N/A"
                # element1 = SportsOracles().SpotracBackup1(str, team)
        print("First: " + element1)
        return element1
    # def Spotrac(self, str, team):
    #     print("1")
    #     salary = "N/A"
    #     df = pd.read_excel('spotrac1.xlsx', engine = 'openpyxl')
        
    #     # find the salary for the given name and team
    #     try:
    #         # result = df[(df['name'].str.find(name) != -1) & (df['team'] == team)]
    #         wb = openpyxl.load_workbook('spotrac1.xlsx')
    #         sheet = wb['Sheet1']
    
    #         for row in sheet.iter_rows(values_only=True):
    #             if str in row[0] and team in row[1]:
    #                 salary = row[2]
    #         print(salary)
    #     except:
    #         print("1 failed")
    #         salary = SportsOracles().SpotracBackup(str, team)
    #     return salary
    def SpotracList(self, website):
        result = requests.get(website) #website being scraped
        print(result.status_code)
        # print(result.headers)
        src = result.content
        soup = BeautifulSoup(src, 'lxml')

        salaries = [] #array of salaries which is what is being searched below

        for td_tag in soup.find_all('td'): # I've seen simpler versions of this but I couldn't get them to work for the data I needed
            span_tag = td_tag.find('span')
            try:
                if 'class' in span_tag.attrs:
                    temp = str(span_tag)
                    temp2 = temp.translate ({ord(c): "" for c in ' !@#%^&*()[]{};:./<>?\|`~-=_+classeminfotitlespanvoidVFrhRLT"'}) #getting rid of all the unnecessary characters
                    salaries.append(temp2)

            except:
                pass
        names = [] 
        for h3_tag in soup.find_all('h3'):
            a_tag = h3_tag.find('a')
            try:
                if 'style' in a_tag.attrs:
                    temp = str(a_tag)
                    temp2 = temp.translate ({ord(c): "" for c in ',!@#%^&*()[]{};:./<>?\|`~=+"1234567890'}) #gets rid of characters
                    temp3 = temp2.replace('a classteam-name hrefhttpswwwspotraccomredirectplayer styleline-height px','') #gets rid of strings
                    temp3 = temp3[:-1] #gets rid of an extra character always attached to the end of the name
                    names.append(temp3)

            except:
                pass   
        print(names)
        print(salaries)
        return names, salaries
    # def PFR(self, str, team):
    #     driver = webdriver.Chrome('./chromedriver')
    #     url = 'https://www.google.com/'
    #     driver.get(url)
    #     element1 = driver.find_element(By.NAME, "q")
    #     search = "PFR " + str + " " + team
    #     element1.send_keys(search)
    #     element1.send_keys(Keys.RETURN)
    #     link = driver.find_elements(By.CSS_SELECTOR, "div#rso > div:nth-of-type(1) > .Ww4FFb.g.tF2Cxc.vt6azd  .UK95Uc.Z26q7c.jGGQ5e .DKV0Md.LC20lb.MBeuO")
    #     if not link:
    #         # print("NOT FOUND **")
    #         element5 = "N/A"
    #         element6 = "N/A"
    #     else:
    #         element2 = driver.find_element(By.CSS_SELECTOR, "div#rso > div:nth-of-type(1) > .Ww4FFb.g.tF2Cxc.vt6azd  .UK95Uc.Z26q7c.jGGQ5e .DKV0Md.LC20lb.MBeuO").click()
    #         driver.maximize_window()
    #         try:
    #             PFFOvrLink = driver.find_element(By.CSS_SELECTOR, "table#passing > tbody > tr:nth-of-type(6) > td:nth-of-type(31)")
    #             element6 = PFFOvrLink.text
    #             # print("PFF: " + element6)
    #         except NoSuchElementException:
    #             # print("Overall Not Found")
    #             element6 = "N/A"
    #         print("success")
    #     return element5, element6

    def PFR(self, str, team):

        driver = webdriver.Chrome('./chromedriver')
        url = 'https://www.google.com/'
        driver.get(url)
        element1 = driver.find_element(By.NAME, "q")
        search = "PFR " + str + " " + team
        element1.send_keys(search)
        element1.send_keys(Keys.RETURN)
        link = driver.find_elements(By.CSS_SELECTOR, "div#rso > div:nth-of-type(1) > .Ww4FFb.g.tF2Cxc.vt6azd  .UK95Uc.Z26q7c.jGGQ5e .DKV0Md.LC20lb.MBeuO")
        if not link:
            element6 = "N/A"
        else:
            element2 = driver.find_element(By.CSS_SELECTOR, "div#rso > div:nth-of-type(1) > .Ww4FFb.g.tF2Cxc.vt6azd  .UK95Uc.Z26q7c.jGGQ5e .DKV0Md.LC20lb.MBeuO").click()
            driver.maximize_window()
            try:
                PFROvrLink = driver.find_elements(By.CSS_SELECTOR, '[data-stat="av"]')
                length = len(PFROvrLink)
                # for element in PFFOvrLink:
                #     print(element.text)
                try:
                    element6 = PFROvrLink[length-2].text
                except:
                    element6 = "N/A"
            except NoSuchElementException:
                element6 = "N/A"
        # print(element6)
        return element6
    def PositionsSal(self, str):
        str1 = str.replace(" ", "")
        positionsSalaries = {"QB": "$5,740,484", "HB": "$1,771,731", "FB": "$1,266,088", 
        "WR": "$2,390,355", "TE": "$1,983,261", "OT": "$3,536,357", "OG": "$2,451,945", 
        "C": "$2,558,048", "DE": "$3,198,393", "DT": "$2,991,040", "OLB": "$2,782,424", 
        "MLB": "$3,350,875", "CB": "$2,206,489", "FS": "$3,548,345", "SS": "$3,237,867", 
        "K": "$1,933,002", "P": "$1,517,049", "LS": "$949,060"}
        salary = positionsSalaries.get(str1, "N/A")
        # print(salary)
        return salary
    def ExcelSheet(self, length, positions, names, positionsSal, age, madden, PFF, salaries, team,  PFR):
        workbook = xlsxwriter.Workbook(team + '.xlsx') #the rest is making an excel sheet
        worksheet = workbook.add_worksheet()

        # Widen the first column to make the text clearer.
        worksheet.set_column('A:A', 12.5)
        worksheet.set_column('B:B', 7.5)
        worksheet.set_column('C:C', 19.5)
        worksheet.set_column('D:D', 13.5)
        worksheet.set_column('E:E', 7.5)
        worksheet.set_column('F:F', 7.5)
        worksheet.set_column('G:G', 7.5)
        worksheet.set_column('H:H', 6.5)
        worksheet.set_column('I:I', 7.5)
        worksheet.set_column('J:J', 8.5)
        worksheet.set_column('K:K', 9.5)
        worksheet.set_column('L:L', 15.5)
        worksheet.set_column('M:M', 21.5)
        worksheet.set_column('N:N', 18.5)
        worksheet.set_column('O:O', 16.5)
        worksheet.set_column('P:P', 8.5)
        worksheet.set_column('Q:Q', 11.5)
        worksheet.set_column('R:R', 10.5)
        worksheet.set_column('S:S', 8.5)
        worksheet.set_column('T:T', 12.5)
        worksheet.set_column('U:U', 12.5)
        # Add a bold format to use to highlight cells.
        format = workbook.add_format({'align': 'center'})
        format.set_font_name("Times New Roman")
        format.set_font_size(12)
        format.set_align('center')
        format.set_align('vcenter')

        worksheet.write('A1', 'Team', format)
        worksheet.write('B1', 'Position', format)
        worksheet.write('C1', 'Player', format)
        worksheet.write('D1', 'Average Salary', format)
        worksheet.write('E1', 'Age', format)
        worksheet.write('F1', 'Madden', format)
        worksheet.write('G1', 'PFF', format)
        worksheet.write('H1', 'PFR', format)
        worksheet.write('I1', 'Lineups', format)
        worksheet.write('J1', 'Injury Prj', format)
        worksheet.write('K1', 'Overall', format)
        worksheet.write('L1', 'Ovr (adj. Age/Inj)', format)
        worksheet.write('M1', 'Average Position Salary', format)
        worksheet.write('N1', 'Deserved Salary', format)
        worksheet.write('O1', 'Salary Difference', format)
        worksheet.write('P1', '', format)
        worksheet.write('Q1', 'Yrs w/ Team', format)
        worksheet.write('R1', 'Acquired', format)
        worksheet.write('S1', 'Drafted', format)
        worksheet.write('T1', 'Draft Class', format)
        worksheet.write('U1', 'Position Rank', format)
        # Write some numbers, with row/column notation.
        for i in range(length):
            j = str(i+2)
            worksheet.write(i+1, 0, team, format)
            worksheet.write(i+1, 1, positions[i], format)
            worksheet.write(i+1, 2, names[i], format)
            worksheet.write(i+1, 4, age[i], format)
            worksheet.write(i+1, 5, madden[i], format)
            worksheet.write(i+1, 6, PFF[i], format)
            worksheet.write(i+1, 7, PFR[i], format)
            worksheet.write(i+1, 3, salaries[i], format)
            worksheet.write(i+1, 10, "=(F" + j + "*0.3611)+(G" + j + "*1.1*0.3611)+(H" + j + "*5.2*0.0666)+(I" + j + "*0.2111)", format) #some formulas I came up with for the values that I want
            worksheet.write(i+1, 11, "=(K"  + j + ")*((100+(29-E"  + j + "))/100)-(J" + j + "*0.9)", format)
            worksheet.write(i+1, 12, positionsSal[i], format)
            worksheet.write(i+1, 13, "=(M"  + j + "*((L"  + j + "-63)/5.45))+M" + j, format)
            worksheet.write(i+1, 14, "=(N" + j + "-D" + j + ")", format)
            worksheet.write(i+1, 15, '=IF(O' + j + '<(-3000000), "Overpaid", "")', format)

        workbook.close()
    def RunLoops(teamSite, teamSpot, team):
        ARIPlayers = []
        ARIPFFRatings = []
        ARIPlayerAge = []
        ARIPlayerAvSal = []
        ARIPosition = []
        ARIMadden = []
        ARISpotracPlayers = []
        ARISpotracSalaries = []
        ARIPFRRatings = []
        # ARISalaries = []
        for temp in SportsOracles().names(teamSite):
            print(temp)
            time1 = time.perf_counter()
            # age, overall = SportsOracles().PFF(temp, team)
            PFRoverall = SportsOracles().PFR(temp, team)
            time2 = time.perf_counter()
            time3 = time2-time1
            print(time3, " sec")
            ARIPlayers.append(temp)
            # ARIMadden.append(temp)
            # ARIPlayerAge.append(temp)
            # ARIPlayerAvSal.append(temp)        
            # ARIPFFRatings.append(temp)
            # ARIPosition.append(temp)
            # ARIPFFRatings.append(overall)
            # ARIPlayerAge.append(age)
            ARIPFFRatings.append("0")
            ARIPlayerAge.append("0")
            ARIPFRRatings.append(PFRoverall)
        ARISpotracPlayers, ARISpotracSalaries = SportsOracles().SpotracList(teamSpot)
        ARISalaries = [0 for i in range(len(ARIPlayers))]
        counter1 = 0
        for temp3 in ARISpotracPlayers:
            counter2 = 0
            # trans1 = temp3.translate ({ord(c): "" for c in ",:.`'1234567890-"})
            # trans2 = temp3.translate ({ord(c): " " for c in ",:.`'1234567890-"})
            trans3 = temp3.translate ({ord(c): "" for c in ",:.`'1234567890-"}).translate ({ord(c): "" for c in " "}).replace("Jr", "").replace("II", "").replace("III", "").replace("IV", "").lower()

            for temp5 in range(len(ARIPlayers)):
                temp6 = ARIPlayers[counter2]
                transMad1 = temp6.translate ({ord(c): "" for c in ",:.`'1234567890-"}).translate ({ord(c): "" for c in " "}).replace("Jr", "").replace("II", "").replace("III", "").replace("IV", "").lower()
                if temp3 in ARIPlayers[counter2] or trans3 in temp6:
                    ARISalaries[counter2] = ARISpotracSalaries[counter1]
                counter2 += 1
            counter1 += 1
        counter3 = 0
        for temp6 in ARISalaries:
            print(counter3)
            if temp6 == 0:
                ARISalaries[counter3] = SportsOracles().SpotracBackup(ARIPlayers[counter3], team)
            counter3 += 1
        for temp1 in SportsOracles().Position(teamSite):
            AvSal = SportsOracles().PositionsSal(temp1)
            ARIPlayerAvSal.append(AvSal)
            # print(temp1)
            ARIPosition.append(temp1)
        for temp2 in SportsOracles().madden(teamSite):
            # print(temp2)
            ARIMadden.append(temp2)
        length1 = len(ARIPlayers)
        length2 = len(ARIPFFRatings)
        length3 = len(ARIPlayerAge)
        length4 = len(ARIPosition)
        length5 = len(ARIPlayerAvSal)
        length6 = len(ARIMadden)
        length7 = len(ARISalaries)
        length8 = len(ARISpotracPlayers)
        length9 = len(ARISpotracSalaries)
        print(ARIPlayers)
        print(ARIPFFRatings)
        print(ARIPlayerAge)
        print(ARIPosition)
        print(ARIPlayerAvSal) 
        print(ARIMadden)
        print(ARISalaries)
        print(ARISpotracPlayers)
        print(ARISpotracSalaries)
        # print(length1 + " " + length2  + " " + 
        # length3  + " " + length4  + " " + length5  + " " + length6)
        SportsOracles().ExcelSheet(length1, ARIPosition, ARIPlayers, ARIPlayerAvSal, 
        ARIPlayerAge, ARIMadden, ARIPFFRatings, ARISalaries, team, ARIPFRRatings)
def main():
    team1 = "https://www.maddenratings.com/teams/arizona-cardinals"
    teamSpot1 = "https://www.spotrac.com/nfl/rankings/average/arizona-cardinals/"
    team2 = "https://www.maddenratings.com/teams/atlanta-falcons"
    teamSpot2 = "https://www.spotrac.com/nfl/rankings/average/atlanta-falcons/"
    team3 = "https://www.maddenratings.com/teams/baltimore-ravens"
    teamSpot3 = "https://www.spotrac.com/nfl/rankings/average/baltimore-ravens/"
    team4 = "https://www.maddenratings.com/teams/buffalo-bills"
    teamSpot4 = "https://www.spotrac.com/nfl/rankings/average/buffalo-bills/"
    team5 = "https://www.maddenratings.com/teams/carolina-panthers"
    teamSpot5 = "https://www.spotrac.com/nfl/rankings/average/carolina-panthers/"
    team6 = "https://www.maddenratings.com/teams/chicago-bears"
    teamSpot6 = "https://www.spotrac.com/nfl/rankings/average/chicago-bears/"
    team7 = "https://www.maddenratings.com/teams/cincinnati-bengals"
    teamSpot7 = "https://www.spotrac.com/nfl/rankings/average/cincinnati-bengals/"
    team8 = "https://www.maddenratings.com/teams/cleveland-browns"
    teamSpot8 = "https://www.spotrac.com/nfl/rankings/average/cleveland-browns/"
    team9 = "https://www.maddenratings.com/teams/dallas-cowboys"
    teamSpot9 = "https://www.spotrac.com/nfl/rankings/average/dallas-cowboys/"
    team10 = "https://www.maddenratings.com/teams/denver-broncos"
    teamSpot10 = "https://www.spotrac.com/nfl/rankings/average/denver-broncos/"
    team11 = "https://www.maddenratings.com/teams/detroit-lions"
    teamSpot11 = "https://www.spotrac.com/nfl/rankings/average/detroit-lions/"
    team12 = "https://www.maddenratings.com/teams/green-bay-packers"
    teamSpot12 = "https://www.spotrac.com/nfl/rankings/average/green-bay-packers/"
    team13 = "https://www.maddenratings.com/teams/houston-texans"
    teamSpot13 = "https://www.spotrac.com/nfl/rankings/average/houston-texans/"
    team14 = "https://www.maddenratings.com/teams/indianapolis-colts"
    teamSpot14 = "https://www.spotrac.com/nfl/rankings/average/indianapolis-colts/"
    team15 = "https://www.maddenratings.com/teams/jacksonville-jaguars"
    teamSpot15 = "https://www.spotrac.com/nfl/rankings/average/jacksonville-jaguars/"
    team16 = "https://www.maddenratings.com/teams/kansas-city-chiefs"
    teamSpot16 = "https://www.spotrac.com/nfl/rankings/average/kansas-city-chiefs/"
    team17 = "https://www.maddenratings.com/teams/las-vegas-raiders"
    teamSpot17 = "https://www.spotrac.com/nfl/rankings/average/las-vegas-raiders/"
    team18 = "https://www.maddenratings.com/teams/los-angeles-chargers"
    teamSpot18 = "https://www.spotrac.com/nfl/rankings/average/los-angeles-chargers/"
    team19 = "https://www.maddenratings.com/teams/los-angeles-rams"
    teamSpot19 = "https://www.spotrac.com/nfl/rankings/average/los-angeles-rams/"
    team20 = "https://www.maddenratings.com/teams/miami-dolphins"
    teamSpot20 = "https://www.spotrac.com/nfl/rankings/average/miami-dolphins/"
    team21 = "https://www.maddenratings.com/teams/minnesota-vikings"
    teamSpot21 = "https://www.spotrac.com/nfl/rankings/average/minnesota-vikings/"
    team22 = "https://www.maddenratings.com/teams/new-england-patriots"
    teamSpot22 = "https://www.spotrac.com/nfl/rankings/average/new-england-patriots/"
    team23 = "https://www.maddenratings.com/teams/new-orleans-saints"
    teamSpot23 = "https://www.spotrac.com/nfl/rankings/average/new-orleans-saints/"
    team24 = "https://www.maddenratings.com/teams/new-york-giants"
    teamSpot24 = "https://www.spotrac.com/nfl/rankings/average/new-york-giants/"
    team25 = "https://www.maddenratings.com/teams/new-york-jets"
    teamSpot25 = "https://www.spotrac.com/nfl/rankings/average/new-york-jets/"
    team26 = "https://www.maddenratings.com/teams/philadelphia-eagles"
    teamSpot26 = "https://www.spotrac.com/nfl/rankings/average/philadelphia-eagles/"
    team27 = "https://www.maddenratings.com/teams/pittsburgh-steelers"
    teamSpot27 = "https://www.spotrac.com/nfl/rankings/average/pittsburgh-steelers/"
    team28 = "https://www.maddenratings.com/teams/san-francisco-49ers"
    teamSpot28 = "https://www.spotrac.com/nfl/rankings/average/san-francisco-49ers/"
    team29 = "https://www.maddenratings.com/teams/seattle-seahawks"
    teamSpot29 = "https://www.spotrac.com/nfl/rankings/average/seattle-seahawks/"
    team30 = "https://www.maddenratings.com/teams/tampa-bay-buccaneers"
    teamSpot30 = "https://www.spotrac.com/nfl/rankings/average/tampa-bay-buccaneers/"
    team31 = "https://www.maddenratings.com/teams/tennessee-titans"
    teamSpot31 = "https://www.spotrac.com/nfl/rankings/average/tennessee-titans/"
    team32 = "https://www.maddenratings.com/teams/washington-commanders"
    teamSpot32 = "https://www.spotrac.com/nfl/rankings/average/washington-commanders/"
    
    ARI = "ARI"
    ATL = 'ATL'
    BAL = 'BAL'
    BUF = 'BUF'
    CAR = "CAR"
    CHI = "CHI"
    CIN = "CIN"
    CLE = "CLE"
    DAL = "DAL"
    DEN = 'DEN'
    DET = 'DET'
    GB = 'GB'
    HOU = 'HOU'
    IND = 'IND'
    JAC = 'JAC'
    KC = 'KC'
    LV = 'LV'
    LAC = 'LAC'
    LAR = 'LAR'
    MIA = 'MIA'
    MIN = 'MIN'
    NE = 'NE'
    NO = 'NO'
    NYG = 'NYG'
    NYJ = 'NYJ'
    PHI = 'PHI'
    PIT = 'PIT'
    SF = 'SF'
    SEA = 'SEA'
    TB = 'TB'
    TEN = 'TEN'
    WAS = 'WAS'
    SportsOracles.RunLoops(team1, teamSpot1, ARI)
    SportsOracles.RunLoops(team2, teamSpot2, ATL)
    SportsOracles.RunLoops(team3, teamSpot3, BAL)
    SportsOracles.RunLoops(team4, teamSpot4, BUF)
    SportsOracles.RunLoops(team5, teamSpot5, CAR)
    SportsOracles.RunLoops(team6, teamSpot6, CHI)
    SportsOracles.RunLoops(team7, teamSpot7, CIN)
    SportsOracles.RunLoops(team8, teamSpot8, CLE)
    SportsOracles.RunLoops(team9, teamSpot9, DAL)
    SportsOracles.RunLoops(team10, teamSpot10, DEN)
    SportsOracles.RunLoops(team11, teamSpot11, DET)
    SportsOracles.RunLoops(team12, teamSpot12, GB)
    SportsOracles.RunLoops(team13, teamSpot13, HOU)
    SportsOracles.RunLoops(team14, teamSpot14, IND)
    SportsOracles.RunLoops(team15, teamSpot15, JAC)
    SportsOracles.RunLoops(team16, teamSpot16, KC)
    SportsOracles.RunLoops(team17, teamSpot17, LV)
    SportsOracles.RunLoops(team18, teamSpot18, LAC)
    SportsOracles.RunLoops(team19, teamSpot19, LAR)
    SportsOracles.RunLoops(team20, teamSpot20, MIA)
    SportsOracles.RunLoops(team21, teamSpot21, MIN)
    SportsOracles.RunLoops(team22, teamSpot22, NE)
    SportsOracles.RunLoops(team23, teamSpot23, NO)
    SportsOracles.RunLoops(team24, teamSpot24, NYG)
    SportsOracles.RunLoops(team25, teamSpot25, NYJ)
    SportsOracles.RunLoops(team26, teamSpot26, PHI)
    SportsOracles.RunLoops(team27, teamSpot27, PIT)
    SportsOracles.RunLoops(team28, teamSpot28, SF)
    SportsOracles.RunLoops(team29, teamSpot29, SEA)
    SportsOracles.RunLoops(team30, teamSpot30, TB)
    SportsOracles.RunLoops(team31, teamSpot31, TEN)
    SportsOracles.RunLoops(team32, teamSpot32, WAS)

    # ARIPlayers = []
    # ARIPFFRatings = []
    # ARIPlayerAge = []
    # ARIPlayerAvSal = []
    # ARIPosition = []
    # ARIMadden = []
    # ARISpotracPlayers = []
    # ARISpotracSalaries = []
    # # ARISalaries = []
    # for temp in SportsOracles().names(team1):
    #     print(temp)
    #     time1 = time.perf_counter()
    #     age, overall = SportsOracles().PFF(temp, ARI)
    #     time2 = time.perf_counter()
    #     time3 = time2-time1
    #     print(time3, " sec")
    #     ARIPlayers.append(temp)
    #     # ARIMadden.append(temp)
    #     # ARIPlayerAge.append(temp)
    #     # ARIPlayerAvSal.append(temp)        
    #     # ARIPFFRatings.append(temp)
    #     # ARIPosition.append(temp)
    #     ARIPFFRatings.append(overall)
    #     ARIPlayerAge.append(age)
    # ARISpotracPlayers, ARISpotracSalaries = SportsOracles().SpotracList(teamSpot1)
    # ARISalaries = [0 for i in range(len(ARIPlayers))]
    # counter1 = 0
    # for temp3 in ARISpotracPlayers:
    #     counter2 = 0
    #     # trans1 = temp3.translate ({ord(c): "" for c in ",:.`'1234567890-"})
    #     # trans2 = temp3.translate ({ord(c): " " for c in ",:.`'1234567890-"})
    #     trans3 = temp3.translate ({ord(c): "" for c in ",:.`'1234567890-"}).translate ({ord(c): "" for c in " "}).replace("Jr", "").replace("II", "").replace("III", "").replace("IV", "").lower()

    #     for temp5 in range(len(ARIPlayers)):
    #         temp6 = ARIPlayers[counter2]
    #         transMad1 = temp6.translate ({ord(c): "" for c in ",:.`'1234567890-"}).translate ({ord(c): "" for c in " "}).replace("Jr", "").replace("II", "").replace("III", "").replace("IV", "").lower()
    #         if temp3 in ARIPlayers[counter2] or trans3 in temp6:
    #             ARISalaries[counter2] = ARISpotracSalaries[counter1]
    #         counter2 += 1
    #     counter1 += 1
    # counter3 = 0
    # for temp6 in ARISalaries:
    #     print(counter3)
    #     if temp6 == 0:
    #         ARISalaries[counter3] = SportsOracles().SpotracBackup(ARIPlayers[counter3], ARI)
    #     counter3 += 1
    # for temp1 in SportsOracles().Position(team1):
    #     AvSal = SportsOracles().PositionsSal(temp1)
    #     ARIPlayerAvSal.append(AvSal)
    #     # print(temp1)
    #     ARIPosition.append(temp1)
    # for temp2 in SportsOracles().madden(team1):
    #     # print(temp2)
    #     ARIMadden.append(temp2)
    # length1 = len(ARIPlayers)
    # length2 = len(ARIPFFRatings)
    # length3 = len(ARIPlayerAge)
    # length4 = len(ARIPosition)
    # length5 = len(ARIPlayerAvSal)
    # length6 = len(ARIMadden)
    # length7 = len(ARISalaries)
    # length8 = len(ARISpotracPlayers)
    # length9 = len(ARISpotracSalaries)
    # print(ARIPlayers)
    # print(ARIPFFRatings)
    # print(ARIPlayerAge)
    # print(ARIPosition)
    # print(ARIPlayerAvSal) 
    # print(ARIMadden)
    # print(ARISalaries)
    # print(ARISpotracPlayers)
    # print(ARISpotracSalaries)
    # # print(length1 + " " + length2  + " " + 
    # # length3  + " " + length4  + " " + length5  + " " + length6)
    # SportsOracles().ExcelSheet(length1, ARIPosition, ARIPlayers, ARIPlayerAvSal, 
    # ARIPlayerAge, ARIMadden, ARIPFFRatings, ARISalaries, ARI)
if __name__ == "__main__":
    main()
